"""
Convert 6 separate category xlsx files -> data/products/*.json
Source: D:/Ai Auto Flow/shopee_affiliate_products/<thai-name>.xlsx
Each file: single sheet, header row 1, data rows 2-21 (Top 20)
Columns: 1=rank 2=name 3=image_url 4=price 5=original_price 6=discount_%
         7=sold 8=rating 9=shopee_product_url 10=shopee_affiliate_url
         11=lazada_affiliate_url 12=commission_percent
"""
import json, re, sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

SRC_DIR = Path(r"D:/Ai Auto Flow/shopee_affiliate_products")
JSON_DIR = Path(__file__).parent / "data" / "products"

# Thai filename -> slug
FILE_MAP = {
    "สินค้าขายดี": "bestsellers",
    "เครื่องใช้ในบ้าน": "home-goods",
    "เสื้อผ้าแฟชั่นผู้หญิง": "women-fashion",
    "สัตว์เลี้ยง": "pets",
    "อาหารและเครื่องดื่ม": "food-drinks",
    "เครื่องใช้ไฟฟ้าภายในบ้าน": "home-appliances",
}

SCORE_DEFAULT = {1: 9.8, 2: 9.5, 3: 9.2, 4: 8.9, 5: 8.6, 6: 8.3,
                 7: 8.0, 8: 7.7, 9: 7.4, 10: 7.1}


def score_for(rank: int) -> float:
    if rank in SCORE_DEFAULT:
        return SCORE_DEFAULT[rank]
    return round(max(7.0 - (rank - 10) * 0.15, 5.0), 1)


def clean_name(name: str) -> str:
    name = re.sub(r'[^ -⻿　-�]', '', name)
    name = re.sub(r'(?i)(พร้อมส่ง|ส่งด่วน|ของแท้|ราคาถูก|ลดราคา|โปรโมชั่น|แท้100%|แท้💯|\d+%off)[!!\s]*', '', name)
    for sep in ['|', '(', '[']:
        if sep in name:
            name = name.split(sep)[0]
    name = ' '.join(name.split()).strip()
    if len(name) > 40:
        name = name[:40].rsplit(' ', 1)[0]
    return name or "สินค้า"


def clean_sold(sold: str) -> str:
    sold = sold.strip()
    sold = re.sub(r'^ขายได้\s*', '', sold)
    sold = re.sub(r'\s*ชิ้น$', '', sold)
    return sold.strip() or "-"


def v(ws, row, col):
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""


def num(x, default=0):
    try:
        return int(float(x)) if x not in ("", "None") else default
    except Exception:
        return default


total = 0
for thai_name, slug in FILE_MAP.items():
    path = SRC_DIR / f"{thai_name}.xlsx"
    if not path.exists():
        print(f"MISSING: {path}")
        continue
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    products = []

    for row in range(2, ws.max_row + 1):
        rank = num(v(ws, row, 1))
        if not rank:
            continue
        name = clean_name(str(v(ws, row, 2)).strip())
        image = str(v(ws, row, 3)).strip()
        price = num(v(ws, row, 4))
        orig = num(v(ws, row, 5), price)
        disc = num(v(ws, row, 6), round((1 - price / orig) * 100) if orig > price > 0 else 0)
        sold = clean_sold(str(v(ws, row, 7)))
        try:
            rating = round(float(v(ws, row, 8)), 1) if str(v(ws, row, 8)) not in ("", "None") else 4.8
        except Exception:
            rating = 4.8
        affiliate_url = str(v(ws, row, 10)).strip()
        product_url = str(v(ws, row, 9)).strip()
        lazada_url = str(v(ws, row, 11)).strip()

        shopee_url = affiliate_url if affiliate_url and affiliate_url != "None" else product_url
        if not image or image in ("None", ""):
            image = f"https://placehold.co/400x400/e8f5e9/2d6a4f?text=rank{rank}"

        products.append({
            "rank": rank,
            "name": name or f"สินค้าอันดับ {rank}",
            "image": image,
            "price": price,
            "originalPrice": orig,
            "discount": disc,
            "sold": sold,
            "rating": rating,
            "score": score_for(rank),
            "shopeeUrl": shopee_url or "https://shopee.co.th",
            "lazadaUrl": lazada_url if lazada_url and lazada_url != "None" else "",
            "highlight": "ขายดีที่สุด" if rank == 1 else "",
            "reason": "",
            "pros": ["คุณภาพดี", "ราคาคุ้มค่า", "ส่งเร็ว"],
        })

    products.sort(key=lambda p: p["rank"])
    out = JSON_DIR / f"{slug}.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    total += len(products)
    first = products[0]["name"][:30] if products else "empty"
    print(f"[{slug}] {len(products)} products — {first}")

print(f"\nDone. {total} products total")
