from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

categories = [
    ("bestsellers", "สินค้าขายดีประจำเดือนนี้"),
    ("home-goods", "เครื่องใช้ในบ้าน"),
    ("women-fashion", "เสื้อผ้าแฟชั่นผู้หญิง"),
    ("pets", "สัตว์เลี้ยง"),
    ("food-drinks", "อาหารและเครื่องดื่ม"),
    ("home-appliances", "เครื่องใช้ไฟฟ้าภายในบ้าน"),
]

# Columns: user fills = yellow bg, auto-filled = green bg, rank = blue
headers = [
    ("rank",               6,  "auto"),
    ("shopee_product_url", 55, "user"),
    ("affiliate_url",      55, "user"),
    ("name",               45, "auto"),
    ("image_url",          50, "auto"),
    ("price",              10, "auto"),
    ("original_price",     14, "auto"),
    ("discount_%",         12, "auto"),
    ("sold",               10, "auto"),
    ("rating",             8,  "auto"),
    ("score",              8,  "user"),
    ("highlight",          20, "user"),
    ("reason",             50, "user"),
    ("pro_1",              22, "user"),
    ("pro_2",              22, "user"),
    ("pro_3",              22, "user"),
]

header_fill   = PatternFill("solid", start_color="1B4332")
header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
user_fill     = PatternFill("solid", start_color="FFF9C4")   # yellow = user input
auto_fill     = PatternFill("solid", start_color="E8F5E9")   # green  = auto scraped
rank_fill     = PatternFill("solid", start_color="D8F3DC")
thin          = Side(style="thin", color="BBBBBB")
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

for slug, label in categories:
    ws = wb.create_sheet(title=slug)

    # Title
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Top 10 — {label}  |  🟡 กรอกเอง  |  🟢 auto-fill จาก script"
    ws["A1"].font = Font(bold=True, color="1B4332", name="Arial", size=11)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row
    for col, (h, w, kind) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 30

    # Data rows
    for r in range(1, 11):
        row = r + 2
        for col, (_, _, kind) in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value="")
            if col == 1:
                cell.value = r
                cell.fill = rank_fill
                cell.font = Font(bold=True, name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = user_fill if kind == "user" else auto_fill
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 42

    ws.freeze_panes = "B3"

wb.save("D:\\Projects\\shopee-ranking\\products_template.xlsx")
print("Template saved")
