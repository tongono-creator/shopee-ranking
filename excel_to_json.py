"""
Convert products_template.xlsx -> data/products/*.json
Excel column layout (Gemini template):
1=rank 2=name 3=image_url 4=price 5=original_price 6=discount_%
7=sold 8=rating 9=shopee_product_url 10=shopee_affiliate_url 11=lazada_affiliate_url
"""
import json, sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = Path(__file__).parent / "products_template.xlsx"
JSON_DIR   = Path(__file__).parent / "data" / "products"

SHEETS = ["bestsellers","home-goods","women-fashion","pets","food-drinks","home-appliances"]
SCORE_DEFAULT = {1:9.8,2:9.5,3:9.2,4:8.9,5:8.6,6:8.3,7:8.0,8:7.7,9:7.4,10:7.1}

def v(ws, row, col):
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""

wb = load_workbook(EXCEL_PATH, data_only=True)

for sheet_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"Skip: {sheet_name}")
        continue
    ws = wb[sheet_name]
    products = []

    for row in range(3, 13):
        rank = int(float(v(ws, row, 1))) if v(ws, row, 1) else 0
        if not rank:
            continue

        name          = str(v(ws, row, 2)).strip()
        image         = str(v(ws, row, 3)).strip()
        price_raw     = v(ws, row, 4)
        orig_raw      = v(ws, row, 5)
        disc_raw      = v(ws, row, 6)
        sold          = str(v(ws, row, 7)).strip()
        rating_raw    = v(ws, row, 8)
        affiliate_url = str(v(ws, row, 10)).strip()
        lazada_url    = str(v(ws, row, 11)).strip()
        product_url   = str(v(ws, row, 9)).strip()

        try: price = int(float(price_raw)) if price_raw else 0
        except: price = 0
        try: orig = int(float(orig_raw)) if orig_raw else price
        except: orig = price
        try: disc = int(float(disc_raw)) if disc_raw else (round((1-price/orig)*100) if orig > price > 0 else 0)
        except: disc = 0
        try: rating = round(float(rating_raw), 1) if rating_raw else 0.0
        except: rating = 0.0

        shopee_url = affiliate_url if affiliate_url and affiliate_url != "None" else product_url

        if not image or image in ("None",""):
            image = f"https://placehold.co/400x400/e8f5e9/2d6a4f?text=rank{rank}"

        products.append({
            "rank": rank,
            "name": name or f"สินค้าอันดับ {rank}",
            "image": image,
            "price": price,
            "originalPrice": orig,
            "discount": disc,
            "sold": sold or "-",
            "rating": rating,
            "score": SCORE_DEFAULT.get(rank, 7.0),
            "shopeeUrl": shopee_url or "https://shopee.co.th",
            "lazadaUrl": lazada_url if lazada_url and lazada_url != "None" else "",
            "highlight": "ขายดีที่สุด" if rank == 1 else "",
            "reason": "",
            "pros": ["คุณภาพดี", "ราคาคุ้มค่า", "ส่งเร็ว"],
        })

    out = JSON_DIR / f"{sheet_name}.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"[{sheet_name}] {len(products)} products — {products[0]['name'][:30] if products else 'empty'}")

print("\nDone")
