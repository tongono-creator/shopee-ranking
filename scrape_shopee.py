"""
Shopee Product Scraper
Usage: python scrape_shopee.py
"""
import asyncio
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

# Fix Windows console encoding
sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = Path(__file__).parent / "products_template.xlsx"

SHEETS = [
    "bestsellers",
    "home-goods",
    "women-fashion",
    "pets",
    "food-drinks",
    "home-appliances",
]

COL = {
    "rank": 1,
    "shopee_product_url": 2,
    "affiliate_url": 3,
    "name": 4,
    "image_url": 5,
    "price": 6,
    "original_price": 7,
    "discount_%": 8,
    "sold": 9,
    "rating": 10,
}


async def scrape_product(page, url: str) -> dict:
    print(f"  Scraping: {url[:80]}...")
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(3500)

        data = await page.evaluate("""() => {
            const getMeta = (sel) => {
                const el = document.querySelector(sel);
                return el ? (el.content || el.getAttribute('content') || '') : '';
            };

            const name = getMeta('meta[property="og:title"]') ||
                         (document.querySelector('h1') || {}).innerText || '';

            const image = getMeta('meta[property="og:image"]');

            let price = 0, originalPrice = 0;
            const priceEls = document.querySelectorAll('[class*="price"]');
            const nums = [];
            priceEls.forEach(el => {
                const txt = el.innerText || '';
                const match = txt.match(/[\\d,]+/g);
                if (match) match.forEach(n => {
                    const v = parseInt(n.replace(/,/g, ''));
                    if (v > 0 && v < 1000000) nums.push(v);
                });
            });
            if (nums.length >= 2) {
                price = Math.min(...nums);
                originalPrice = Math.max(...nums);
            } else if (nums.length === 1) {
                price = nums[0];
                originalPrice = nums[0];
            }

            let rating = 0;
            const ratingEl = document.querySelector('[class*="rating"] [class*="number"]') ||
                             document.querySelector('[class*="shopee-rating"]');
            if (ratingEl) {
                const r = parseFloat(ratingEl.innerText);
                if (!isNaN(r)) rating = r;
            }

            let sold = '';
            const soldEl = document.querySelector('[class*="sold"]');
            if (soldEl) {
                const txt = (soldEl.innerText || '').trim();
                const m = txt.match(/[\\d,k]+/i);
                if (m) sold = m[0] + '+';
            }

            return { name, image, price, originalPrice, sold, rating };
        }""")

        return data
    except Exception as e:
        print(f"  Error: {e}")
        return {}


async def main():
    wb = load_workbook(EXCEL_PATH)
    changed = False

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="th-TH",
        )
        page = await context.new_page()

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            print(f"\n[{sheet_name}]")

            for row in range(3, 13):
                url_cell = ws.cell(row=row, column=COL["shopee_product_url"])
                name_cell = ws.cell(row=row, column=COL["name"])
                url = url_cell.value

                if not url or not str(url).startswith("http"):
                    continue
                if name_cell.value and str(name_cell.value).strip():
                    print(f"  Row {row}: already filled, skip")
                    continue

                data = await scrape_product(page, str(url))
                if not data:
                    continue

                if data.get("name"):
                    ws.cell(row=row, column=COL["name"]).value = data["name"]
                if data.get("image"):
                    ws.cell(row=row, column=COL["image_url"]).value = data["image"]
                if data.get("price"):
                    ws.cell(row=row, column=COL["price"]).value = data["price"]
                if data.get("originalPrice") and data["originalPrice"] != data.get("price"):
                    ws.cell(row=row, column=COL["original_price"]).value = data["originalPrice"]
                    if data["price"] and data["originalPrice"] > data["price"]:
                        disc = round((1 - data["price"] / data["originalPrice"]) * 100)
                        ws.cell(row=row, column=COL["discount_%"]).value = disc
                if data.get("sold"):
                    ws.cell(row=row, column=COL["sold"]).value = data["sold"]
                if data.get("rating"):
                    ws.cell(row=row, column=COL["rating"]).value = data["rating"]

                rank = ws.cell(row=row, column=COL["rank"]).value
                print(f"  Rank {rank}: done")
                changed = True
                await asyncio.sleep(2)

        await browser.close()

    if changed:
        wb.save(EXCEL_PATH)
        print(f"\nSaved: {EXCEL_PATH}")
    else:
        print("\nNothing to update — check URLs in Excel")


if __name__ == "__main__":
    asyncio.run(main())
