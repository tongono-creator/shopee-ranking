"""
Kalodata Shopee Product Scraper Wizard
Usage: python scrape_kalodata.py
"""
import asyncio
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

# Fix Windows console encoding for Thai characters
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = Path("D:/Projects/shopee-ranking/products_template.xlsx")

SHEETS = [
    ("1", "bestsellers", "สินค้าขายดี"),
    ("2", "home-goods", "เครื่องใช้ในบ้าน"),
    ("3", "women-fashion", "เสื้อผ้าแฟชั่นผู้หญิง"),
    ("4", "pets", "สัตว์เลี้ยง"),
    ("5", "food-drinks", "อาหารและเครื่องดื่ม"),
    ("6", "home-appliances", "เครื่องใช้ไฟฟ้าภายในบ้าน"),
]

COLUMNS = {
    "rank": 1,
    "name": 2,
    "image_url": 3,
    "price": 4,
    "original_price": 5,
    "discount_%": 6,
    "sold": 7,
    "rating": 8,
    "shopee_product_url": 9,
    "shopee_affiliate_url": 10,
    "lazada_affiliate_url": 11,
}

async def scrape_active_page(page) -> list:
    print("\n[Scraper] Extracting table data from Kalodata active page...")
    
    # Evaluate JavaScript in the page to extract products dynamically
    products = await page.evaluate("""() => {
        // Find rows that look like product rows
        const rows = Array.from(document.querySelectorAll('tr, [role="row"], .ant-table-row, div[class*="row"]'));
        
        const productRows = rows.filter(row => {
            const images = row.querySelectorAll('img');
            // Product rows should have product images and a decent amount of text
            if (images.length === 0) return false;
            if (row.innerText.length < 30) return false;
            
            // Filter out headers or footer cells
            if (row.innerText.includes('หมวดหมู่') && row.innerText.includes('รายได้')) return false;
            return true;
        });

        console.log('Detected product rows count:', productRows.length);
        
        return productRows.slice(0, 10).map((row, i) => {
            // Find images
            const imgs = Array.from(row.querySelectorAll('img')).map(img => img.src).filter(src => {
                return src && !src.includes('avatar') && !src.includes('logo') && !src.includes('icon');
            });
            const image = imgs[0] || '';

            // Find links
            const links = Array.from(row.querySelectorAll('a')).map(a => a.href);
            let shopeeUrl = links.find(href => href.includes('shopee.co.th')) || '';
            if (!shopeeUrl) {
                const productLink = links.find(href => href.includes('/product/'));
                if (productLink) shopeeUrl = productLink;
            }

            // Split text blocks to identify name, price, sold
            const texts = row.innerText.split('\\n').map(t => t.trim()).filter(Boolean);
            
            // Rank is usually first text block if it is a single number, or just i + 1
            let rank = i + 1;
            if (texts[0] && /^\\d+$/.test(texts[0])) {
                rank = parseInt(texts[0]);
            }

            // Product name is usually the longest text block (excluding money or stats)
            let name = '';
            const potentialNames = texts.filter(t => t.length > 15 && !t.includes('฿') && !t.includes('%') && !t.includes('+'));
            if (potentialNames.length > 0) {
                name = potentialNames[0];
            } else {
                name = texts.find(t => t.length > 8) || 'สินค้า';
            }

            // Price
            let price = 0;
            const priceText = texts.find(t => t.includes('฿'));
            if (priceText) {
                const m = priceText.match(/[\\d,]+/);
                if (m) price = parseInt(m[0].replace(/,/g, ''));
            }

            // Sold
            let sold = '';
            const soldText = texts.find(t => t.includes('+') || t.toLowerCase().includes('k') || t.includes('ชิ้น') || t.includes('ล้าน'));
            if (soldText) {
                sold = soldText;
            } else {
                const idx = texts.indexOf(name);
                if (idx !== -1 && texts[idx + 1]) sold = texts[idx + 1];
            }

            return { rank, name, image, price, sold, shopeeUrl };
        });
    }""")
    
    return products

async def main():
    print("=" * 60)
    print("      KALODATA SHOPEE PRODUCT SCRAPER WIZARD")
    print("=" * 60)
    print("This script will open a browser for you to log in to Kalodata")
    print("and navigate to the category/product ranking page.")
    print("Then it will automatically capture the products and save them to Excel.")
    print("-" * 60)
    
    async with async_playwright() as p:
        # Launch headed browser
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="th-TH",
            viewport={'width': 1366, 'height': 768}
        )
        page = await context.new_page()
        
        print("\n[System] Opening Kalodata.com...")
        await page.goto("https://www.kalodata.com/", wait_until="domcontentloaded")
        
        print("\n" + "!" * 50)
        print("ACTION REQUIRED:")
        print("1. Log in to your Kalodata account in the browser window.")
        print("2. Navigate to the Shopee Category or Product leaderboard you want to scrape.")
        print("3. Ensure the product table is visible on the screen.")
        print("!" * 50)
        
        input("\n>>> After you have navigated to the correct page on Kalodata, press [Enter] here to start scraping...")
        
        products = await scrape_active_page(page)
        
        if not products:
            print("\n[Error] No products could be detected on the active page.")
            print("Please make sure you are on the product list/table page and try again.")
            await browser.close()
            return
            
        print(f"\n[Scraper] Successfully detected {len(products)} products:")
        for p_item in products:
            print(f"  #{p_item['rank']}: {p_item['name'][:40]}... | Price: ฿{p_item['price']} | Sold: {p_item['sold']}")
            
        print("\n" + "=" * 50)
        print("Select the target Category Sheet to save to:")
        for num, slug, name in SHEETS:
            print(f"  [{num}] {name} ({slug})")
        print("=" * 50)
        
        sheet_choice = input("\nEnter sheet number (1-6) or press Enter to skip saving: ").strip()
        
        selected_sheet = None
        for num, slug, name in SHEETS:
            if sheet_choice == num:
                selected_sheet = slug
                break
                
        if not selected_sheet:
            print("\n[System] Cancelled saving. Exiting...")
            await browser.close()
            return
            
        print(f"\n[System] Saving products to Excel sheet: '{selected_sheet}'...")
        
        wb = load_workbook(EXCEL_PATH)
        if selected_sheet not in wb.sheetnames:
            print(f"[Error] Sheet '{selected_sheet}' not found in Excel template!")
            await browser.close()
            return
            
        ws = wb[selected_sheet]
        
        # Write to rows 3 to 12 (Top 10 products)
        for idx, item in enumerate(products[:10]):
            row = 3 + idx
            
            # Fill Excel columns based on our mappings
            ws.cell(row=row, column=COLUMNS["rank"], value=item["rank"])
            ws.cell(row=row, column=COLUMNS["name"], value=item["name"])
            
            if item["image"]:
                ws.cell(row=row, column=COLUMNS["image_url"], value=item["image"])
                
            if item["price"]:
                ws.cell(row=row, column=COLUMNS["price"], value=item["price"])
                # Fallback original price to the same price if unknown
                ws.cell(row=row, column=COLUMNS["original_price"], value=item["price"])
                ws.cell(row=row, column=COLUMNS["discount_%"], value=0)
                
            if item["sold"]:
                # Clean sold count (remove labels if any, keep number + k or similar)
                ws.cell(row=row, column=COLUMNS["sold"], value=item["sold"])
                
            ws.cell(row=row, column=COLUMNS["rating"], value=4.8) # default rating
            
            if item["shopeeUrl"]:
                ws.cell(row=row, column=COLUMNS["shopee_product_url"], value=item["shopeeUrl"])
                
            print(f"  Saved Rank {item['rank']} -> Row {row}")
            
        wb.save(EXCEL_PATH)
        print(f"\n[Success] Excel template updated successfully at {EXCEL_PATH}!")
        print("You can now run:")
        print("  python excel_to_json.py")
        print("to compile this data to JSON for your website.")
        
        await browser.close()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nExiting...")
