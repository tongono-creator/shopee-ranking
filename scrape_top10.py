"""
Shopee Top 10 Scraper
Usage: python scrape_top10.py
"""
import asyncio
import sys
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

# Fix Windows console encoding for Thai characters
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = Path("D:/Projects/shopee-ranking/products_template.xlsx")

CATEGORIES = [
    ("bestsellers", "https://shopee.co.th/search?keyword=%E0%B8%AA%E0%B8%B4%E0%B8%99%E0%B8%84%E0%B9%82%E0%B8%B2%E0%B8%82%E0%B8%B2%E0%B8%A2%E0%B8%94%E0%B8%B5&sortBy=sales"),
    ("home-goods", "https://shopee.co.th/search?keyword=%E0%B9%80%E0%B8%84%E0%B8%A3%E0%B8%B7%E0%B9%88%E0%B8%AD%E0%B8%87%E0%B9%83%E0%B8%8A%E0%B9%89%E0%B9%83%E0%B8%99%E0%B8%9A%E0%B9%89%E0%B8%B2%E0%B8%99&sortBy=sales"),
    ("women-fashion", "https://shopee.co.th/search?keyword=%E0%B9%80%E0%B8%81%E0%B8%B7%E0%B9%89%E0%B8%AD%E0%B8%9C%E0%B9%89%E0%B8%B2%E0%B9%81%E0%B8%9F%E0%B8%84%E0%B8%B1%E0%B9%88%E0%B8%99%E0%B8%9C%E0%B8%B9%E0%B9%89%E0%B8%AB%E0%B8%8D%E0%B8%B4%E0%B8%87&sortBy=sales"),
    ("pets", "https://shopee.co.th/search?keyword=%E0%B8%AA%E0%B8%B1%E0%B8%95%E0%B8%A7%E0%B9%8C%E0%B9%80%E0%B8%A5%E0%B8%B5%E0%B9%89%E0%B8%A2%E0%B8%87&sortBy=sales"),
    ("food-drinks", "https://shopee.co.th/search?keyword=%E0%B8%AD%E0%B8%B2%E0%B8%AB%E0%B8%B2%E0%B8%A3%E0%B9%81%E0%B8%A5%E0%B8%B0%E0%B9%80%E0%B8%84%E0%B8%A3%E0%B8%B7%E0%B9%88%E0%B8%AD%E0%B8%87%E0%B8%94%E0%B8%B7%E0%B9%88%E0%B8%A1&sortBy=sales"),
    ("home-appliances", "https://shopee.co.th/search?keyword=%E0%B9%80%E0%B8%84%E0%B8%A3%E0%B8%B7%E0%B9%88%E0%B8%AD%E0%B8%87%E0%B9%83%E0%B8%8A%E0%B9%89%E0%B9%84%E0%B8%9F%E0%B8%9F%E0%B9%89%E0%B8%B2%E0%B8%A0%E0%B8%B2%E0%B8%A2%E0%B9%83%E0%B8%99%E0%B8%9A%E0%B9%89%E0%B8%B2%E0%B8%99&sortBy=sales"),
]

COLUMNS = [
    "rank", "name", "image_url", "price", "original_price", "discount_%", 
    "sold", "rating", "shopee_product_url", "shopee_affiliate_url", "lazada_affiliate_url"
]

async def scrape_category(page, url, sheet_name):
    print(f"\nScraping category: {sheet_name}\nURL: {url}")
    try:
        await page.goto(url, wait_until="networkidle", timeout=60000)
    except Exception as e:
        print(f"  Error loading page: {e}. Trying again with domcontentloaded...")
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except:
            print(f"  Failed to load {url}")
            return []
    
    # Handle language modal (Click "ไทย" if present)
    try:
        thai_button = page.get_by_role("button", name="ไทย")
        if await thai_button.is_visible(timeout=5000):
            await thai_button.click()
            await page.wait_for_timeout(2000)
    except:
        pass

    # Scroll down gradually to trigger lazy loading of images and list items
    for _ in range(8):
        await page.mouse.wheel(0, 800)
        await page.wait_for_timeout(1000)
    
    print(f"  Page title: {await page.title()}")
    
    if "ตรวจสอบ" in await page.title() or "Verification" in await page.title():
        print("  Blocked by verification!")
        
    await page.screenshot(path=f"debug_{sheet_name}.png")
    html = await page.content()
    with open(f"debug_{sheet_name}.html", "w", encoding="utf-8") as f:
        f.write(html[:50000]) # First 50k chars
        
    products = await page.evaluate("""() => {
        // Find product cards. The selectors cover both normal search results and the "top products" layout
        let items = Array.from(document.querySelectorAll('.shopee-search-item-result__item, [data-sqe="item"], .top-products-section__item, .row > div'));
        
        // Log found items for debugging
        console.log('Found items count:', items.length);
        
        if (items.length === 0) {
            return [];
        }
        
        return items.slice(0, 10).map((el, i) => {
            // ... (rest of mapping)
            const getEl = (sel) => el.querySelector(sel);
            const getText = (sel) => {
                const element = getEl(sel);
                return element ? element.innerText.trim() : "";
            };
            
            // Link & Image
            const linkEl = el.querySelector('a');
            let url = "";
            if (linkEl) {
                let href = linkEl.getAttribute('href');
                if (href && !href.startsWith('http')) {
                    url = "https://shopee.co.th" + href.split('?')[0];
                } else if (href) {
                    url = href.split('?')[0];
                }
            }
            
            const imgEl = el.querySelector('img');
            let img = "";
            if (imgEl) {
                img = imgEl.getAttribute('src') || imgEl.getAttribute('data-src') || "";
                if (img.startsWith('//')) img = 'https:' + img;
            }
            
            // Name
            let name = getText('[data-sqe="name"], .shopee-item-card__name, [class*="name"], .yQmmFK');
            if (!name) {
                const nameEl = el.querySelector('.top-products-item__name') || el.querySelector('div[class*="name"]');
                if (nameEl) name = nameEl.innerText.trim();
            }
            
            // Price
            let price = "0";
            const priceText = getText('[class*="price"], ._32697j, .bPCQZk');
            const priceNums = priceText.match(/[\\d,]+/g);
            if (priceNums && priceNums.length > 0) {
                price = priceNums[0].replace(/,/g, '');
            }
            
            // Original Price
            let originalPrice = "";
            const origText = getText('[class*="original-price"], ._29_7S9, .old-price');
            const origNums = origText.match(/[\\d,]+/g);
            if (origNums && origNums.length > 0) {
                originalPrice = origNums[0].replace(/,/g, '');
            }
            
            // Discount
            let discount = "";
            const discText = getText('[class*="discount"], ._2p6968');
            if (discText) {
                const match = discText.match(/(\\d+)/);
                if (match) discount = match[1];
            }
            
            // Sold Count
            let sold = "0";
            const soldText = getText('[class*="sold"], ._184768, .shopee-item-card__play-count, .top-products-item__sold');
            const soldMatch = soldText.match(/([\\d,\\.]+[kKmM]?)/);
            if (soldMatch) {
                sold = soldMatch[1].toLowerCase();
            }
            
            // Rating
            let rating = "5.0";
            const ratingEl = el.querySelector('.shopee-rating-stars__active');
            if (ratingEl) {
                const style = ratingEl.getAttribute('style');
                const match = style.match(/width:\\s*([\\d\\.]+)%/);
                if (match) {
                    rating = (parseFloat(match[1]) / 20).toFixed(1);
                }
            }
            
            return {
                "rank": i + 1,
                "name": name,
                "image_url": img,
                "price": price,
                "original_price": originalPrice,
                "discount_%": discount,
                "sold": sold,
                "rating": rating,
                "shopee_product_url": url,
                "shopee_affiliate_url": "",
                "lazada_affiliate_url": ""
            };
        });
    }""")
    
    if not products:
        print(f"  Warning: No products found for {sheet_name}")
        
    for p in products:
        print(f"  #{p['rank']}: {p['name'][:40]}... | ฿{p['price']} | Sold: {p['sold']}")
        
    return products

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="th-TH",
            viewport={'width': 1280, 'height': 800}
        )
        page = await context.new_page()
        await Stealth().apply_stealth_async(page)
        
        print("Initial visit to homepage to establish session...")
        await page.goto("https://shopee.co.th/", wait_until="networkidle", timeout=60000)
        await asyncio.sleep(5)
        
        all_data = {}
        for sheet_name, url in CATEGORIES:
            all_data[sheet_name] = await scrape_category(page, url, sheet_name)
            await asyncio.sleep(3)
            
        await browser.close()
        
        wb = Workbook()
        wb.remove(wb.active)
        
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_fill = PatternFill("solid", start_color="1B4332")
        border = Border(left=Side(style='thin', color='BBBBBB'), 
                        right=Side(style='thin', color='BBBBBB'), 
                        top=Side(style='thin', color='BBBBBB'), 
                        bottom=Side(style='thin', color='BBBBBB'))
        
        for sheet_name, _ in CATEGORIES:
            products = all_data.get(sheet_name, [])
            ws = wb.create_sheet(title=sheet_name)
            
            # Write Headers
            for col, header in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                
                width = 15
                if header in ["name", "image_url", "shopee_product_url"]:
                    width = 45
                elif header == "rank":
                    width = 8
                ws.column_dimensions[get_column_letter(col)].width = width
                
            # Write Data
            for row_idx, product in enumerate(products, 2):
                for col_idx, header in enumerate(COLUMNS, 1):
                    val = product.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    
                    if header == "rank":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        
            ws.row_dimensions[1].height = 25
            
        os.makedirs(EXCEL_PATH.parent, exist_ok=True)
        wb.save(EXCEL_PATH)
        print(f"\\n[Success] Excel saved to {EXCEL_PATH}")

if __name__ == "__main__":
    asyncio.run(main())
